import os
import concurrent.futures
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import json
import google.generativeai as genai
import httpx
import base64
import typing_extensions as typing
import time
import mimetypes
from urllib.request import urlopen
import re
import logging
import csv
from dotenv import load_dotenv
load_dotenv()
import threading
import time
from collections import deque

# === Constants ===
IMAGE_FORMATS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
MAX_PROCESSED_ROWS = 520
# OUTPUT_FILE = "processed_output.csv" # Not used
INPUT_DIR = "parallel_input_split_1_files"
OUTPUT_DIR = "parallel_output_split_1_files"
FINAL_OUTPUT_FILE = os.path.join(OUTPUT_DIR, "merged_output.csv")

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] [%(threadName)s] %(message)s",
    handlers=[
        logging.FileHandler(f"{OUTPUT_DIR}/processing.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ‚úÖ Add this below logger setup
def validate_csv(input_file, logger):
    bad_rows = []
    with open(input_file, newline='') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            try:
                pass  # You can add your row-level validation here
            except Exception as e:
                logger.warning(f"Malformed row at line {i+1}: {e}")
                bad_rows.append(i+1)
    return bad_rows

def get_gemini_tokens_from_env():
    tokens = []
    for key in os.environ:
        if key.startswith("GEMINI_TOKEN"):
            tokens.append(os.environ[key])
    if not tokens:
        logging.error("[Gemini] No Gemini tokens found in environment variables!")
    return tokens

GEMINI_TOKENS = get_gemini_tokens_from_env()

current_token_index = 0

def get_next_gemini_token():
    global current_token_index
    if current_token_index < len(GEMINI_TOKENS):
        token = GEMINI_TOKENS[current_token_index]
        logging.info(f"[Gemini] Using token: -----") #{token}
        return token
    return None

def switch_to_next_token():
    global current_token_index
    current_token_index += 1
    if current_token_index >= len(GEMINI_TOKENS):
        logging.error("[Gemini] All tokens exhausted!")
        return None
    token = get_next_gemini_token()
    if token:
        genai.configure(api_key=token)
        global model
        model = genai.GenerativeModel(
            model_name="gemini-2.0-flash",
            generation_config={
                "response_mime_type": "application/json",
                "response_schema": AnalysisResponse,
            },
        )
        return token
    return None

# === Gemini Model Setup ===
class AnalysisResponse(typing.TypedDict):
    answers: list[str]
    reasonings: list[str]

initial_token = get_next_gemini_token()
if not initial_token:
    raise ValueError("[Gemini] No valid Gemini tokens found!")

genai.configure(api_key=initial_token)
model = genai.GenerativeModel(
    model_name="gemini-2.0-flash",
    generation_config={
        "response_mime_type": "application/json",
        "response_schema": AnalysisResponse,
    },
)

# === Utility functions ===
def calculate_relevance_tag(answers):
    if not answers or not isinstance(answers, list):
        return 'Irrelevant'
    yes_count = sum(1 for answer in answers if str(answer).upper() == 'YES')
    total_count = len(answers)
    percentage = (yes_count / total_count) * 100 if total_count > 0 else 0
    if percentage >= 50:
        return 'Relevant'
    elif percentage > 0:
        return 'Partially Relevant'
    else:
        return 'Irrelevant'

def adjust_excel_formatting(output_file):
    # This function is for .xlsx, but the script now saves .csv
    # It won't be called by the current logic but is harmless to keep.
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(output_file)
    except Exception as e:
        logging.warning(f"Could not apply Excel formatting to {output_file}: {e}")


def get_image_as_base64(url: str) -> str:
    with urlopen(url) as response:
        image_data = response.read()
    mime_type, _ = mimetypes.guess_type(url)
    if not mime_type:
        mime_type = "image/jpeg"
    base64_data = base64.b64encode(image_data).decode("utf-8")
    return f"data:{mime_type};base64,{base64_data}"

# Track timestamps of recent requests
_request_times = deque()
_request_lock = threading.Lock()
MAX_REQUESTS_PER_MINUTE = 2000

def rate_limiter():
    """Block until we are under the 2000 req/min limit."""
    global _request_times
    with _request_lock:
        now = time.time()
        # Remove requests older than 60 seconds
        while _request_times and now - _request_times[0] > 60:
            _request_times.popleft()

        if len(_request_times) >= MAX_REQUESTS_PER_MINUTE:
            sleep_time = 60 - (now - _request_times[0])
            if sleep_time > 0:
                logging.info(f"[RateLimiter] Throttling for {sleep_time:.2f} seconds to stay under 2000 req/min...")
                time.sleep(sleep_time)
                return rate_limiter()  # Recheck after sleep

        _request_times.append(time.time())


def process_image(task_evidence_link, task_evidence_question, max_retries=3):
    global current_token_index
    retries = 0
    while retries < max_retries:
        try:
            rate_limiter()
            image = httpx.get(task_evidence_link)
            prompt = f"""You are an educational evidence validator. Analyse the given image... {task_evidence_question}"""
            response = model.generate_content([
                {"mime_type": "image/jpeg", "data": base64.b64encode(image.content).decode("utf-8")},
                prompt,
            ])
            response_json = json.loads(response.text)
            return response_json
        except Exception as e:
            error_str = str(e).lower()
            if any(k in error_str for k in ["rate limit", "quota", "429", "resource_exhausted"]):
                logging.warning("[Gemini] Rate limit or quota exceeded. Switching token...")
                if switch_to_next_token():
                    continue
                else:
                    logging.warning("[Gemini] No more tokens. Retrying in 60 seconds...")
                    time.sleep(60)
                    retries += 1
            else:
                logging.error(f"[Gemini] Error: {e}")
                retries += 1
    logging.error("[Gemini] Max retries reached.")
    return {"error": "Max retries reached"}


# === Main processing ===
def main(input_file, worker_id=None):
    # ‚úÖ --- Stats variables ---
    api_calls = 0
    api_successes = 0
    api_failures = 0
    success_list = []
    failed_list = []
    # --- End stats ---

    try:
        logging.info(f"[Worker {worker_id}] Starting processing for {input_file}")

        if not os.path.exists(input_file):
            logging.error(f"[Worker {worker_id}] File not found: {input_file}")
            return None

        df = pd.read_excel(input_file) if input_file.endswith(".xlsx") else pd.read_csv(input_file)
        df_filtered = df[
            ~df["Task Evidence"].isin([None, "Null"])
            & ~df["Task Evidence Question"].isin([None, "Null"])
        ].dropna(subset=["Task Evidence", "Task Evidence Question"])

        processed_count = 0
        task_evidence_qa = []
        task_evidence_qa_reason = []
        relevance_tags = []

        for idx, row in df_filtered.iterrows():
            task_evidence = str(row["Task Evidence"]).strip()
            task_question = str(row["Task Evidence Question"]).strip()

            if any(task_evidence.lower().endswith(ext) for ext in IMAGE_FORMATS):
                logging.info(f"[Worker {worker_id}] Processing image row {idx+1}/{len(df_filtered)}")
                
                api_calls += 1  # ‚úÖ Track API call attempt
                response = process_image(task_evidence, task_question)
                
                if isinstance(response, dict) and "answers" in response and "reasonings" in response:
                    answers = response["answers"]
                    reasonings = response["reasonings"]
                    task_evidence_qa.append(answers)
                    task_evidence_qa_reason.append(reasonings)
                    relevance_tags.append(calculate_relevance_tag(answers))
                    
                    api_successes += 1 # ‚úÖ Track success
                    success_list.append(task_evidence) # ‚úÖ Add to success list
                else:
                    logging.warning(f"[Worker {worker_id}] Invalid response at row {idx+1}")
                    task_evidence_qa.append(None)
                    task_evidence_qa_reason.append(None)
                    relevance_tags.append('Irrelevant')
                    
                    api_failures += 1 # ‚úÖ Track failure
                    failed_list.append(task_evidence) # ‚úÖ Add to failed list
            else:
                logging.info(f"[Worker {worker_id}] Skipping non-image row {idx+1}")
                task_evidence_qa.append(None)
                task_evidence_qa_reason.append(None)
                relevance_tags.append('Irrelevant')

            processed_count += 1
            if processed_count >= MAX_PROCESSED_ROWS:
                logging.info(f"[Worker {worker_id}] Reached max processed rows ({MAX_PROCESSED_ROWS})")
                break

        df_filtered = df_filtered.head(processed_count)
        df_filtered["Task evidence Q and A"] = task_evidence_qa
        df_filtered["Task evidence Q and A Reason"] = task_evidence_qa_reason
        df_filtered["Relevance Tag"] = relevance_tags
        
        df_filtered["Image Preview"] = df_filtered["Task Evidence"].apply(
            lambda x: str(x) if str(x).lower().endswith(tuple(IMAGE_FORMATS)) else ""
        )

        output_filename = os.path.join(OUTPUT_DIR, f"processed_{os.path.basename(input_file).split('.')[0]}.csv")
        df_filtered.to_csv(output_filename, index=False)
        
        logging.info(f"[Worker {worker_id}] Finished processing {input_file}. Output: {output_filename}")

        # ‚úÖ Return the dictionary of stats
        return {
            "output_file": output_filename,
            "rows_attempted": processed_count,
            "api_calls": api_calls,
            "api_successes": api_successes,
            "api_failures": api_failures,
            "success_list": success_list,
            "failed_list": failed_list
        }

    except Exception as e:
        logging.exception(f"[Worker {worker_id}] Failed to process {input_file}: {e}")
        return None


def process_file_parallel(file_path, worker_id):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_stats = main(file_path, worker_id)  # ‚úÖ Get stats dictionary
    if output_stats:
        logging.info(f"[Worker {worker_id}] Output saved as {output_stats['output_file']}")
    else:
        logging.warning(f"[Worker {worker_id}] Processing failed for {file_path}")
    return output_stats  # ‚úÖ Return the entire stats dictionary (or None)

# === Entry point ===
if __name__ == "__main__":
    input_files = [
        os.path.join(INPUT_DIR, file)
        for file in os.listdir(INPUT_DIR)
        if file.endswith((".xlsx", ".csv"))
    ]

    logging.info(f"[Main] Found {len(input_files)} input files to process.")

    # ‚úÖ --- Global Stats Aggregators ---
    total_rows_processed_all = 0
    total_api_calls_all = 0
    total_api_success_all = 0
    total_api_failure_all = 0
    all_success_lists = []
    all_failed_lists = []
    processed_files = [] # List of successful output file paths
    failed_files = [] # List of input files that failed to process
    # --- End Aggregators ---

    with concurrent.futures.ThreadPoolExecutor(max_workers=len(input_files)) as executor:
        futures = {
            executor.submit(process_file_parallel, f, idx + 1): f
            for idx, f in enumerate(input_files)
        }
        for future in concurrent.futures.as_completed(futures):
            original_file = futures[future]
            result_stats = future.result()
            
            if result_stats:  # ‚úÖ Check if processing was successful
                processed_files.append(result_stats["output_file"])
                total_rows_processed_all += result_stats["rows_attempted"]
                total_api_calls_all += result_stats["api_calls"]
                total_api_success_all += result_stats["api_successes"]
                total_api_failure_all += result_stats["api_failures"]
                all_success_lists.extend(result_stats["success_list"])
                all_failed_lists.extend(result_stats["failed_list"])
                logging.info(f"[Main] Worker finished processing: {original_file}")
            else:
                logging.warning(f"[Main] File {original_file} failed to process.")
                failed_files.append(original_file)

    if not processed_files:
        logging.error("[Main] No files processed successfully. Exiting.")
        # ‚úÖ Still log the summary even if exiting
    else:
        try:
            logging.info(f"[Main] Merging {len(processed_files)} files into {FINAL_OUTPUT_FILE}")
            merged_df = pd.concat([pd.read_csv(f) for f in processed_files], ignore_index=True)
            merged_df.to_csv(FINAL_OUTPUT_FILE, index=False)
            logging.info(f"‚úÖ All files processed and merged into: {FINAL_OUTPUT_FILE}")
        except Exception as e:
            logging.exception(f"[Main] Error during merging: {e}")

    # ‚úÖ --- Log the Final Summary ---
    try:
        logging.info("="*80)
        logging.info("===== üöÄ PROCESSING RUN SUMMARY =====")
        logging.info("="*80)
        
        logging.info(f"Total Rows Processed (sum of attempts): {total_rows_processed_all}")
        logging.info(f"Total API Calls (image rows attempted): {total_api_calls_all}")
        logging.info(f"  - ‚úÖ Success: {total_api_success_all}")
        logging.info(f"  - ‚ùå Failed: {total_api_failure_all}")
        
        logging.info("")
        logging.info(f"Total Input Files Processed Successfully: {len(processed_files)}")
        logging.info(f"Total Input Files Failed to Process: {len(failed_files)}")
        if failed_files:
            logging.warning("Failed Input Files:")
            for f in failed_files:
                logging.warning(f"  - {f}")

        logging.info("")
        if all_failed_lists:
            logging.warning(f"List of Failed API Calls ({len(all_failed_lists)}):")
            for item in all_failed_lists:
                logging.warning(f"  - {item}")
        else:
            logging.info("‚úÖ No API call failures recorded.")

        if all_success_lists:
            logging.info(f"List of Successful API Calls ({len(all_success_lists)}):")
            for item in all_success_lists:
                logging.info(f"  - {item}")
        else:
            logging.info("No API call successes recorded.")
            
        logging.info("="*80)
        logging.info("===== üèÅ END OF SUMMARY =====")
        logging.info("="*80 + "\n")
    except Exception as e:
        logging.exception(f"[Main] Failed to write summary to log: {e}")





# import os
# import concurrent.futures
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment
# import json
# import google.generativeai as genai
# import httpx
# import base64
# import typing_extensions as typing
# import time
# import mimetypes
# from urllib.request import urlopen
# import re
# import logging
# import csv
# from dotenv import load_dotenv
# load_dotenv()
# import threading
# import time
# from collections import deque

# # === Constants ===
# IMAGE_FORMATS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
# MAX_PROCESSED_ROWS = 1
# # OUTPUT_FILE = "processed_output.csv"
# INPUT_DIR = "parallel_input_split_files"
# OUTPUT_DIR = "parallel_output_split_files"
# FINAL_OUTPUT_FILE = os.path.join(OUTPUT_DIR, "merged_output.csv")

# # Create output directory if it doesn't exist
# os.makedirs(OUTPUT_DIR, exist_ok=True)

# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] [%(threadName)s] %(message)s",
#     handlers=[
#         logging.FileHandler("parallel_output_split_files/processing.log"),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)

# # ‚úÖ Add this below logger setup
# def validate_csv(input_file, logger):
#     bad_rows = []
#     with open(input_file, newline='') as f:
#         reader = csv.reader(f)
#         for i, row in enumerate(reader):
#             try:
#                 pass  # You can add your row-level validation here
#             except Exception as e:
#                 logger.warning(f"Malformed row at line {i+1}: {e}")
#                 bad_rows.append(i+1)
#     return bad_rows

# def get_gemini_tokens_from_env():
#     tokens = []
#     for key in os.environ:
#         if key.startswith("GEMINI_TOKEN"):
#             tokens.append(os.environ[key])
#     if not tokens:
#         logging.error("[Gemini] No Gemini tokens found in environment variables!")
#     return tokens

# GEMINI_TOKENS = get_gemini_tokens_from_env()

# current_token_index = 0

# def get_next_gemini_token():
#     global current_token_index
#     if current_token_index < len(GEMINI_TOKENS):
#         token = GEMINI_TOKENS[current_token_index]
#         logging.info(f"[Gemini] Using token: -----") #{token}
#         return token
#     return None

# def switch_to_next_token():
#     global current_token_index
#     current_token_index += 1
#     if current_token_index >= len(GEMINI_TOKENS):
#         logging.error("[Gemini] All tokens exhausted!")
#         return None
#     token = get_next_gemini_token()
#     if token:
#         genai.configure(api_key=token)
#         global model
#         model = genai.GenerativeModel(
#             model_name="gemini-2.0-flash",
#             generation_config={
#                 "response_mime_type": "application/json",
#                 "response_schema": AnalysisResponse,
#             },
#         )
#         return token
#     return None

# # === Gemini Model Setup ===
# class AnalysisResponse(typing.TypedDict):
#     answers: list[str]
#     reasonings: list[str]

# initial_token = get_next_gemini_token()
# if not initial_token:
#     raise ValueError("[Gemini] No valid Gemini tokens found!")

# genai.configure(api_key=initial_token)
# model = genai.GenerativeModel(
#     model_name="gemini-2.0-flash",
#     generation_config={
#         "response_mime_type": "application/json",
#         "response_schema": AnalysisResponse,
#     },
# )

# # === Utility functions ===
# def calculate_relevance_tag(answers):
#     if not answers or not isinstance(answers, list):
#         return 'Irrelevant'
#     yes_count = sum(1 for answer in answers if str(answer).upper() == 'YES')
#     total_count = len(answers)
#     percentage = (yes_count / total_count) * 100 if total_count > 0 else 0
#     if percentage >= 50:
#         return 'Relevant'
#     elif percentage > 0:
#         return 'Partially Relevant'
#     else:
#         return 'Irrelevant'

# def adjust_excel_formatting(output_file):
#     wb = load_workbook(output_file)
#     ws = wb.active
#     for row in ws.iter_rows():
#         for cell in row:
#             cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         ws.column_dimensions[col_letter].width = max_length + 2
#     wb.save(output_file)

# def get_image_as_base64(url: str) -> str:
#     with urlopen(url) as response:
#         image_data = response.read()
#     mime_type, _ = mimetypes.guess_type(url)
#     if not mime_type:
#         mime_type = "image/jpeg"
#     base64_data = base64.b64encode(image_data).decode("utf-8")
#     return f"data:{mime_type};base64,{base64_data}"

# # Track timestamps of recent requests
# _request_times = deque()
# _request_lock = threading.Lock()
# MAX_REQUESTS_PER_MINUTE = 2000

# def rate_limiter():
#     """Block until we are under the 2000 req/min limit."""
#     global _request_times
#     with _request_lock:
#         now = time.time()
#         # Remove requests older than 60 seconds
#         while _request_times and now - _request_times[0] > 60:
#             _request_times.popleft()

#         if len(_request_times) >= MAX_REQUESTS_PER_MINUTE:
#             sleep_time = 60 - (now - _request_times[0])
#             if sleep_time > 0:
#                 logging.info(f"[RateLimiter] Throttling for {sleep_time:.2f} seconds to stay under 2000 req/min...")
#                 time.sleep(sleep_time)
#                 return rate_limiter()  # Recheck after sleep

#         _request_times.append(time.time())


# def process_image(task_evidence_link, task_evidence_question, max_retries=3):
#     global current_token_index
#     retries = 0
#     while retries < max_retries:
#         try:
#             rate_limiter()
#             image = httpx.get(task_evidence_link)
#             prompt = f"""You are an educational evidence validator. Analyse the given image... {task_evidence_question}"""
#             response = model.generate_content([
#                 {"mime_type": "image/jpeg", "data": base64.b64encode(image.content).decode("utf-8")},
#                 prompt,
#             ])
#             response_json = json.loads(response.text)
#             return response_json
#         except Exception as e:
#             error_str = str(e).lower()
#             if any(k in error_str for k in ["rate limit", "quota", "429", "resource_exhausted"]):
#                 logging.warning("[Gemini] Rate limit or quota exceeded. Switching token...")
#                 if switch_to_next_token():
#                     continue
#                 else:
#                     logging.warning("[Gemini] No more tokens. Retrying in 60 seconds...")
#                     time.sleep(60)
#                     retries += 1
#             else:
#                 logging.error(f"[Gemini] Error: {e}")
#                 retries += 1
#     logging.error("[Gemini] Max retries reached.")
#     return {"error": "Max retries reached"}


# # === Main processing ===
# def main(input_file, worker_id=None):
#     try:
#         logging.info(f"[Worker {worker_id}] Starting processing for {input_file}")

#         if not os.path.exists(input_file):
#             logging.error(f"[Worker {worker_id}] File not found: {input_file}")
#             return None

#         df = pd.read_excel(input_file) if input_file.endswith(".xlsx") else pd.read_csv(input_file)
#         df_filtered = df[
#             ~df["Task Evidence"].isin([None, "Null"])
#             & ~df["Task Evidence Question"].isin([None, "Null"])
#         ].dropna(subset=["Task Evidence", "Task Evidence Question"])

#         processed_count = 0
#         task_evidence_qa = []
#         task_evidence_qa_reason = []
#         relevance_tags = []

#         for idx, row in df_filtered.iterrows():
#             task_evidence = str(row["Task Evidence"]).strip()
#             task_question = str(row["Task Evidence Question"]).strip()

#             if any(task_evidence.lower().endswith(ext) for ext in IMAGE_FORMATS):
#                 logging.info(f"[Worker {worker_id}] Processing image row {idx+1}/{len(df_filtered)}")
#                 response = process_image(task_evidence, task_question)
#                 if isinstance(response, dict) and "answers" in response and "reasonings" in response:
#                     answers = response["answers"]
#                     reasonings = response["reasonings"]
#                     task_evidence_qa.append(answers)
#                     task_evidence_qa_reason.append(reasonings)
#                     relevance_tags.append(calculate_relevance_tag(answers))
#                 else:
#                     logging.warning(f"[Worker {worker_id}] Invalid response at row {idx+1}")
#                     task_evidence_qa.append(None)
#                     task_evidence_qa_reason.append(None)
#                     relevance_tags.append('Irrelevant')
#             else:
#                 logging.info(f"[Worker {worker_id}] Skipping non-image row {idx+1}")
#                 task_evidence_qa.append(None)
#                 task_evidence_qa_reason.append(None)
#                 relevance_tags.append('Irrelevant')

#             processed_count += 1
#             if processed_count >= MAX_PROCESSED_ROWS:
#                 logging.info(f"[Worker {worker_id}] Reached max processed rows ({MAX_PROCESSED_ROWS})")
#                 break

#         df_filtered = df_filtered.head(processed_count)
#         df_filtered["Task evidence Q and A"] = task_evidence_qa
#         df_filtered["Task evidence Q and A Reason"] = task_evidence_qa_reason
#         df_filtered["Relevance Tag"] = relevance_tags
        
#         # ‚úÖ Remove IMAGE() formula for CSV - it's Excel-specific
#         df_filtered["Image Preview"] = df_filtered["Task Evidence"].apply(
#             lambda x: str(x) if str(x).lower().endswith(tuple(IMAGE_FORMATS)) else ""
#         )

#         # ‚úÖ Changed to save as CSV instead of XLSX
#         output_filename = os.path.join(OUTPUT_DIR, f"processed_{os.path.basename(input_file).split('.')[0]}.csv")
#         df_filtered.to_csv(output_filename, index=False)
        
#         logging.info(f"[Worker {worker_id}] Finished processing {input_file}. Output: {output_filename}")

#         return output_filename  # ‚úÖ Return the output file path

#     except Exception as e:
#         logging.exception(f"[Worker {worker_id}] Failed to process {input_file}: {e}")
#         return None


# def process_file_parallel(file_path, worker_id):
#     os.makedirs(OUTPUT_DIR, exist_ok=True)
#     output_filename = main(file_path, worker_id)  # ‚úÖ Get return from main
#     if output_filename:
#         logging.info(f"[Worker {worker_id}] Output saved as {output_filename}")
#     else:
#         logging.warning(f"[Worker {worker_id}] Processing failed for {file_path}")
#     return output_filename  # ‚úÖ Always return, even if None

# # === Entry point ===
# if __name__ == "__main__":
#     input_files = [
#         os.path.join(INPUT_DIR, file)
#         for file in os.listdir(INPUT_DIR)
#         if file.endswith((".xlsx", ".csv"))
#     ]

#     logging.info(f"[Main] Found {len(input_files)} input files to process.")

#     processed_files = []
#     with concurrent.futures.ThreadPoolExecutor(max_workers=len(input_files)) as executor:
#         futures = {
#             executor.submit(process_file_parallel, f, idx + 1): f
#             for idx, f in enumerate(input_files)
#         }
#         for future in concurrent.futures.as_completed(futures):
#             result_file = future.result()
#             if result_file:  # ‚úÖ Skip failed results
#                 processed_files.append(result_file)
#                 logging.info(f"[Main] Worker finished: {result_file}")
#             else:
#                 logging.warning(f"[Main] A file failed to process.")

#     if not processed_files:
#         logging.error("[Main] No files processed successfully. Exiting.")
#         exit(1)

#     try:
#         logging.info(f"[Main] Merging {len(processed_files)} files into {FINAL_OUTPUT_FILE}")
#         # ‚úÖ Changed to read CSV files instead of Excel files
#         merged_df = pd.concat([pd.read_csv(f) for f in processed_files], ignore_index=True)
#         # ‚úÖ Changed to save merged output as CSV
#         merged_df.to_csv(FINAL_OUTPUT_FILE, index=False)
#         logging.info(f"‚úÖ All files processed and merged into: {FINAL_OUTPUT_FILE}")
#     except Exception as e:
#         logging.exception(f"[Main] Error during merging: {e}")



# import os
# import concurrent.futures
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment
# import json
# import google.generativeai as genai
# import httpx
# import base64
# import typing_extensions as typing
# import time
# import mimetypes
# from urllib.request import urlopen
# import re
# import logging
# import csv
# from dotenv import load_dotenv
# load_dotenv()
# import threading
# import time
# from collections import deque

# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] [%(threadName)s] %(message)s",
#     handlers=[
#         logging.FileHandler("parallel_output_split_files/processing.log"),
#         logging.StreamHandler()
#     ]
# )
# logger = logging.getLogger(__name__)

# # ‚úÖ Add this below logger setup
# def validate_csv(input_file, logger):
#     bad_rows = []
#     with open(input_file, newline='') as f:
#         reader = csv.reader(f)
#         for i, row in enumerate(reader):
#             try:
#                 pass  # You can add your row-level validation here
#             except Exception as e:
#                 logger.warning(f"Malformed row at line {i+1}: {e}")
#                 bad_rows.append(i+1)
#     return bad_rows

# # === Constants ===
# IMAGE_FORMATS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
# MAX_PROCESSED_ROWS = 1
# OUTPUT_FILE = "processed_output.csv"
# INPUT_DIR = "parallel_input_split_files"
# OUTPUT_DIR = "parallel_output_split_files"
# FINAL_OUTPUT_FILE = os.path.join(OUTPUT_DIR, "merged_output.csv")

# def get_gemini_tokens_from_env():
#     tokens = []
#     for key in os.environ:
#         if key.startswith("GEMINI_TOKEN"):
#             tokens.append(os.environ[key])
#     if not tokens:
#         logging.error("[Gemini] No Gemini tokens found in environment variables!")
#     return tokens

# GEMINI_TOKENS = get_gemini_tokens_from_env()

# current_token_index = 0

# def get_next_gemini_token():
#     global current_token_index
#     if current_token_index < len(GEMINI_TOKENS):
#         token = GEMINI_TOKENS[current_token_index]
#         logging.info(f"[Gemini] Using token: -----") #{token}
#         return token
#     return None

# def switch_to_next_token():
#     global current_token_index
#     current_token_index += 1
#     if current_token_index >= len(GEMINI_TOKENS):
#         logging.error("[Gemini] All tokens exhausted!")
#         return None
#     token = get_next_gemini_token()
#     if token:
#         genai.configure(api_key=token)
#         global model
#         model = genai.GenerativeModel(
#             model_name="gemini-2.0-flash",
#             generation_config={
#                 "response_mime_type": "application/json",
#                 "response_schema": AnalysisResponse,
#             },
#         )
#         return token
#     return None

# # === Gemini Model Setup ===
# class AnalysisResponse(typing.TypedDict):
#     answers: list[str]
#     reasonings: list[str]

# initial_token = get_next_gemini_token()
# if not initial_token:
#     raise ValueError("[Gemini] No valid Gemini tokens found!")

# genai.configure(api_key=initial_token)
# model = genai.GenerativeModel(
#     model_name="gemini-2.0-flash",
#     generation_config={
#         "response_mime_type": "application/json",
#         "response_schema": AnalysisResponse,
#     },
# )

# # === Utility functions ===
# def calculate_relevance_tag(answers):
#     if not answers or not isinstance(answers, list):
#         return 'Irrelevant'
#     yes_count = sum(1 for answer in answers if str(answer).upper() == 'YES')
#     total_count = len(answers)
#     percentage = (yes_count / total_count) * 100 if total_count > 0 else 0
#     if percentage >= 50:
#         return 'Relevant'
#     elif percentage > 0:
#         return 'Partially Relevant'
#     else:
#         return 'Irrelevant'

# def adjust_excel_formatting(output_file):
#     wb = load_workbook(output_file)
#     ws = wb.active
#     for row in ws.iter_rows():
#         for cell in row:
#             cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
#     for col in ws.columns:
#         max_length = 0
#         col_letter = col[0].column_letter
#         for cell in col:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         ws.column_dimensions[col_letter].width = max_length + 2
#     wb.save(output_file)

# def get_image_as_base64(url: str) -> str:
#     with urlopen(url) as response:
#         image_data = response.read()
#     mime_type, _ = mimetypes.guess_type(url)
#     if not mime_type:
#         mime_type = "image/jpeg"
#     base64_data = base64.b64encode(image_data).decode("utf-8")
#     return f"data:{mime_type};base64,{base64_data}"

# # Track timestamps of recent requests
# _request_times = deque()
# _request_lock = threading.Lock()
# MAX_REQUESTS_PER_MINUTE = 2000

# def rate_limiter():
#     """Block until we are under the 2000 req/min limit."""
#     global _request_times
#     with _request_lock:
#         now = time.time()
#         # Remove requests older than 60 seconds
#         while _request_times and now - _request_times[0] > 60:
#             _request_times.popleft()

#         if len(_request_times) >= MAX_REQUESTS_PER_MINUTE:
#             sleep_time = 60 - (now - _request_times[0])
#             if sleep_time > 0:
#                 logging.info(f"[RateLimiter] Throttling for {sleep_time:.2f} seconds to stay under 2000 req/min...")
#                 time.sleep(sleep_time)
#                 return rate_limiter()  # Recheck after sleep

#         _request_times.append(time.time())


# def process_image(task_evidence_link, task_evidence_question, max_retries=3):
#     global current_token_index
#     retries = 0
#     while retries < max_retries:
#         try:
#             rate_limiter()
#             image = httpx.get(task_evidence_link)
#             prompt = f"""You are an educational evidence validator. Analyse the given image... {task_evidence_question}"""
#             response = model.generate_content([
#                 {"mime_type": "image/jpeg", "data": base64.b64encode(image.content).decode("utf-8")},
#                 prompt,
#             ])
#             response_json = json.loads(response.text)
#             return response_json
#         except Exception as e:
#             error_str = str(e).lower()
#             if any(k in error_str for k in ["rate limit", "quota", "429", "resource_exhausted"]):
#                 logging.warning("[Gemini] Rate limit or quota exceeded. Switching token...")
#                 if switch_to_next_token():
#                     continue
#                 else:
#                     logging.warning("[Gemini] No more tokens. Retrying in 60 seconds...")
#                     time.sleep(60)
#                     retries += 1
#             else:
#                 logging.error(f"[Gemini] Error: {e}")
#                 retries += 1
#     logging.error("[Gemini] Max retries reached.")
#     return {"error": "Max retries reached"}


# # === Main processing ===
# def main(input_file, worker_id=None):
#     try:
#         logging.info(f"[Worker {worker_id}] Starting processing for {input_file}")

#         if not os.path.exists(input_file):
#             logging.error(f"[Worker {worker_id}] File not found: {input_file}")
#             return None

#         df = pd.read_excel(input_file) if input_file.endswith(".xlsx") else pd.read_csv(input_file)
#         df_filtered = df[
#             ~df["Task Evidence"].isin([None, "Null"])
#             & ~df["Task Evidence Question"].isin([None, "Null"])
#         ].dropna(subset=["Task Evidence", "Task Evidence Question"])

#         processed_count = 0
#         task_evidence_qa = []
#         task_evidence_qa_reason = []
#         relevance_tags = []

#         for idx, row in df_filtered.iterrows():
#             task_evidence = str(row["Task Evidence"]).strip()
#             task_question = str(row["Task Evidence Question"]).strip()

#             if any(task_evidence.lower().endswith(ext) for ext in IMAGE_FORMATS):
#                 logging.info(f"[Worker {worker_id}] Processing image row {idx+1}/{len(df_filtered)}")
#                 response = process_image(task_evidence, task_question)
#                 if isinstance(response, dict) and "answers" in response and "reasonings" in response:
#                     answers = response["answers"]
#                     reasonings = response["reasonings"]
#                     task_evidence_qa.append(answers)
#                     task_evidence_qa_reason.append(reasonings)
#                     relevance_tags.append(calculate_relevance_tag(answers))
#                 else:
#                     logging.warning(f"[Worker {worker_id}] Invalid response at row {idx+1}")
#                     task_evidence_qa.append(None)
#                     task_evidence_qa_reason.append(None)
#                     relevance_tags.append('Irrelevant')
#             else:
#                 logging.info(f"[Worker {worker_id}] Skipping non-image row {idx+1}")
#                 task_evidence_qa.append(None)
#                 task_evidence_qa_reason.append(None)
#                 relevance_tags.append('Irrelevant')

#             processed_count += 1
#             if processed_count >= MAX_PROCESSED_ROWS:
#                 logging.info(f"[Worker {worker_id}] Reached max processed rows ({MAX_PROCESSED_ROWS})")
#                 break

#         df_filtered = df_filtered.head(processed_count)
#         df_filtered["Task evidence Q and A"] = task_evidence_qa
#         df_filtered["Task evidence Q and A Reason"] = task_evidence_qa_reason
#         df_filtered["Relevance Tag"] = relevance_tags
        
#         # ‚úÖ Remove IMAGE() formula for CSV - it's Excel-specific
#         df_filtered["Image Preview"] = df_filtered["Task Evidence"].apply(
#             lambda x: str(x) if str(x).lower().endswith(tuple(IMAGE_FORMATS)) else ""
#         )

#         # ‚úÖ Changed to save as CSV instead of XLSX
#         output_filename = os.path.join(OUTPUT_DIR, f"processed_{os.path.basename(input_file).split('.')[0]}.csv")
#         df_filtered.to_csv(output_filename, index=False)
        
#         logging.info(f"[Worker {worker_id}] Finished processing {input_file}. Output: {output_filename}")

#         return output_filename  # ‚úÖ Return the output file path

#     except Exception as e:
#         logging.exception(f"[Worker {worker_id}] Failed to process {input_file}: {e}")
#         return None


# def process_file_parallel(file_path, worker_id):
#     os.makedirs(OUTPUT_DIR, exist_ok=True)
#     output_filename = main(file_path, worker_id)  # ‚úÖ Get return from main
#     if output_filename:
#         logging.info(f"[Worker {worker_id}] Output saved as {output_filename}")
#     else:
#         logging.warning(f"[Worker {worker_id}] Processing failed for {file_path}")
#     return output_filename  # ‚úÖ Always return, even if None

# # === Entry point ===
# if __name__ == "__main__":
#     input_files = [
#         os.path.join(INPUT_DIR, file)
#         for file in os.listdir(INPUT_DIR)
#         if file.endswith((".xlsx", ".csv"))
#     ]

#     logging.info(f"[Main] Found {len(input_files)} input files to process.")

#     processed_files = []
#     with concurrent.futures.ThreadPoolExecutor(max_workers=len(input_files)) as executor:
#         futures = {
#             executor.submit(process_file_parallel, f, idx + 1): f
#             for idx, f in enumerate(input_files)
#         }
#         for future in concurrent.futures.as_completed(futures):
#             result_file = future.result()
#             if result_file:  # ‚úÖ Skip failed results
#                 processed_files.append(result_file)
#                 logging.info(f"[Main] Worker finished: {result_file}")
#             else:
#                 logging.warning(f"[Main] A file failed to process.")

#     if not processed_files:
#         logging.error("[Main] No files processed successfully. Exiting.")
#         exit(1)

#     try:
#         logging.info(f"[Main] Merging {len(processed_files)} files into {FINAL_OUTPUT_FILE}")
#         # ‚úÖ Changed to read CSV files instead of Excel files
#         merged_df = pd.concat([pd.read_csv(f) for f in processed_files], ignore_index=True)
#         # ‚úÖ Changed to save merged output as CSV
#         merged_df.to_csv(FINAL_OUTPUT_FILE, index=False)
#         logging.info(f"‚úÖ All files processed and merged into: {FINAL_OUTPUT_FILE}")
#     except Exception as e:
#         logging.exception(f"[Main] Error during merging: {e}")